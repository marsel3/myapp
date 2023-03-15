from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, Series, RadarChart
import sqlite3

def datebase():
    connection = sqlite3.connect('server.sqlite')
    cursor = connection.cursor()
    with connection:
        results = cursor.execute("""SELECT student, score FROM results""").fetchall()
    return results

def variance(data):
        # Number of observations
    n = len(data)
         # Mean of the data
    mean = sum(data) / n
        # Square deviations
    deviations = [(x - mean) ** 2 for x in data]
        # Variance
    variance = sum(deviations) / n
    return variance

def MadeExcel():
    wb = Workbook()
    ws = wb.active
    first_row = ('ФИО', 'Баллы')
    ws.append(first_row)
    rows = datebase()

    score_list = [i[1] for i in rows]

    average = sum(score_list) / len(score_list)
    dispersia = variance(score_list)

    gistogram = {}
    for i in score_list:
        if not i in gistogram:
            gistogram[i] = 0
        else:
            gistogram[i] += 1


    ws["M2"] = "Среднее значение"
    ws["N2"] = average
    ws["M3"] = "Дисперсия"
    ws["N3"] = dispersia

    for row in rows:
        ws.append(row)

    print(max([i[1] for i in rows]), print(type(max([i[1] for i in rows]))))
    chart = RadarChart()
    chart.type = "filled"

    labels = Reference(ws, min_col=1, min_row=2, max_row=len(rows)+1)
    data = Reference(ws, min_col=1, max_col=2, min_row=1, max_row=len(rows)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.style = 26
    chart.title = "Результаты тестирования"
    chart.y_axis.delete = True

    ws.add_chart(chart, "C3")

    wb.save("radars.xlsx")



MadeExcel()
